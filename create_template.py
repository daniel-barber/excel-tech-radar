#!/usr/bin/env python3
"""
Create Excel template for Radar Studio import.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Constants
MAX_TEMPLATE_ROWS = 1000

def create_template():
    """Create Excel template with sample data and instructions."""
    wb = Workbook()
    
    # Set document properties
    wb.properties.title = "Radar Template"
    wb.properties.subject = "Template for importing radar projects"
    wb.properties.creator = "Radar Studio"
    
    # ===== DATA SHEET =====
    ws_data = wb.active
    ws_data.title = "Sheet1"
    
    # Headers
    headers = ["name", "ring", "quadrant", "dealSize", "propensityToWin", "isStrategic", "description", "link", "linkName"]
    ws_data.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_num, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Sample data rows
    sample_data = [
        [
            'Cloud Migration Strategy',
            'Current HY',
            'Infrastructure',
            '> $500k',
            'High',
            'Yes',
            '<p>Comprehensive strategy for migrating legacy applications to cloud infrastructure. Includes assessment, planning, and execution phases.</p>',
            'https://example.com/cloud-strategy',
            'View Strategy Doc'
        ],
        [
            'AI-Powered Analytics Platform',
            'Next HY',
            'Data',
            '$100k - $500k',
            'Medium',
            'No',
            '<p>Next-generation analytics platform leveraging <strong>machine learning</strong> and AI to provide predictive insights.</p>',
            'https://example.com/ai-analytics',
            'Platform Overview'
        ],
        [
            'Process Automation Initiative',
            'Year +1',
            'Automation',
            '< $100k',
            'Low',
            'No',
            '<p>Initiative to automate manual business processes using RPA and workflow automation tools.</p>',
            '',
            ''
        ]
    ]
    
    for row_data in sample_data:
        ws_data.append(row_data)
    
    # Add data validation for ring column (column B)
    ring_values = '"Current HY,Next HY,Year +1,Year +2"'
    ring_validation = DataValidation(type="list", formula1=ring_values, allow_blank=False)
    ring_validation.error = 'Please select a valid ring value'
    ring_validation.errorTitle = 'Invalid Ring'
    ws_data.add_data_validation(ring_validation)
    ring_validation.add(f'B2:B{MAX_TEMPLATE_ROWS}')  # Apply to ring column for rows 2-1000
    
    # Add data validation for quadrant column (column C)
    quadrant_values = '"Automation,Data,Infrastructure"'
    quadrant_validation = DataValidation(type="list", formula1=quadrant_values, allow_blank=False)
    quadrant_validation.error = 'Please select a valid quadrant value'
    quadrant_validation.errorTitle = 'Invalid Quadrant'
    ws_data.add_data_validation(quadrant_validation)
    quadrant_validation.add(f'C2:C{MAX_TEMPLATE_ROWS}')  # Apply to quadrant column for rows 2-1000
    
    # Add data validation for dealSize column (column D)
    dealsize_values = '"< $100k,$100k - $500k,> $500k"'
    dealsize_validation = DataValidation(type="list", formula1=dealsize_values, allow_blank=True)
    dealsize_validation.error = 'Please select a valid deal size'
    dealsize_validation.errorTitle = 'Invalid Deal Size'
    ws_data.add_data_validation(dealsize_validation)
    dealsize_validation.add(f'D2:D{MAX_TEMPLATE_ROWS}')  # Apply to dealSize column for rows 2-1000
    
    # Add data validation for propensityToWin column (column E)
    propensity_values = '"Low,Medium,High"'
    propensity_validation = DataValidation(type="list", formula1=propensity_values, allow_blank=True)
    propensity_validation.error = 'Please select a valid propensity to win'
    propensity_validation.errorTitle = 'Invalid Propensity'
    ws_data.add_data_validation(propensity_validation)
    propensity_validation.add(f'E2:E{MAX_TEMPLATE_ROWS}')  # Apply to propensityToWin column for rows 2-1000
    
    # Add data validation for isStrategic column (column F)
    strategic_values = '"Yes,No"'
    strategic_validation = DataValidation(type="list", formula1=strategic_values, allow_blank=True)
    strategic_validation.error = 'Please select Yes or No'
    strategic_validation.errorTitle = 'Invalid Strategic Value'
    ws_data.add_data_validation(strategic_validation)
    strategic_validation.add(f'F2:F{MAX_TEMPLATE_ROWS}')  # Apply to isStrategic column for rows 2-1000
    
    # Auto-adjust column widths
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        if header == "description":
            ws_data.column_dimensions[column_letter].width = 60
        elif header == "name":
            ws_data.column_dimensions[column_letter].width = 30
        elif header in ["link", "linkName"]:
            ws_data.column_dimensions[column_letter].width = 25
        else:
            ws_data.column_dimensions[column_letter].width = 20
    
    # ===== INSTRUCTIONS SHEET =====
    ws_instructions = wb.create_sheet("Instructions")
    
    # Title
    ws_instructions.append(["Radar Studio - Import Template Instructions"])
    ws_instructions["A1"].font = Font(size=16, bold=True, color="4472C4")
    ws_instructions.merge_cells("A1:D1")
    
    ws_instructions.append([])  # Empty row
    
    # Overview
    ws_instructions.append(["Overview"])
    ws_instructions["A3"].font = Font(size=14, bold=True)
    ws_instructions.append(["This template helps you import radar projects into Radar Studio."])
    ws_instructions.append(["Fill in the 'Sheet1' tab with your radar entries, then import the file."])
    ws_instructions.append([])
    
    # Column descriptions
    ws_instructions.append(["Column Descriptions"])
    ws_instructions["A7"].font = Font(size=14, bold=True)
    ws_instructions.append([])
    
    column_descriptions = [
        ["Column", "Required", "Description", "Example"],
        ["name", "Yes", "Unique name for the radar entry", "Cloud Migration Strategy"],
        ["ring", "Yes", "Time horizon: Current HY, Next HY, Year +1, Year +2", "Current HY"],
        ["quadrant", "Yes", "Category: Automation, Data, or Infrastructure", "Infrastructure"],
        ["dealSize", "No", "Deal size category: < $100k, $100k - $500k, > $500k", "> $500k"],
        ["propensityToWin", "No", "Likelihood to win: Low, Medium, High", "High"],
        ["isStrategic", "No", "Strategic importance: Yes or No (shows star indicator)", "Yes"],
        ["description", "No", "HTML description (supports <p>, <strong>, <em>, <a>, <ul>, <li>)", "<p>Detailed description...</p>"],
        ["link", "No", "URL for external reference", "https://example.com/doc"],
        ["linkName", "No", "Display text for the link", "View Documentation"]
    ]
    
    for row_data in column_descriptions:
        ws_instructions.append(row_data)
        row_num = ws_instructions.max_row
        if row_num == 9:  # Header row
            for col_num in range(1, 5):
                cell = ws_instructions.cell(row=row_num, column=col_num)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Column widths for instructions
    ws_instructions.column_dimensions["A"].width = 15
    ws_instructions.column_dimensions["B"].width = 12
    ws_instructions.column_dimensions["C"].width = 60
    ws_instructions.column_dimensions["D"].width = 30
    
    ws_instructions.append([])
    
    # Import steps
    ws_instructions.append(["How to Import"])
    ws_instructions[f"A{ws_instructions.max_row}"].font = Font(size=14, bold=True)
    ws_instructions.append([])
    
    import_steps = [
        "1. Fill in your radar entries in the 'Sheet1' tab",
        "2. Use the dropdown menus for Ring, Quadrant, Deal Size, Propensity, and Strategic columns",
        "3. Name and Ring are required; Quadrant must be one of: Automation, Data, Infrastructure",
        "4. Delete the sample rows (rows 2-4) or replace them with your data",
        "5. Save the file with your desired project name (e.g., 'My Project.xlsx')",
        "6. In Radar Studio, click 'Import' and drag-and-drop your file",
        "7. Your project will be uploaded and appear automatically in the project list"
    ]
    
    for step in import_steps:
        ws_instructions.append([step])
    
    ws_instructions.append([])
    
    # Tips
    ws_instructions.append(["Tips"])
    ws_instructions[f"A{ws_instructions.max_row}"].font = Font(size=14, bold=True)
    ws_instructions.append([])
    
    tips = [
        "• Use HTML in descriptions for rich formatting (bold, italic, links, lists)",
        "• Deal Size affects the circle size on the radar (larger deals = bigger circles)",
        "• Propensity to Win affects the circle color (High=green, Medium=yellow, Low=red)",
        "• Strategic items show a white star in the center of the circle",
        "• Quadrants must be: Automation, Data, or Infrastructure",
        "• Ring values must match: Current HY, Next HY, Year +1, or Year +2",
        "• Leave optional fields empty if not needed"
    ]
    
    for tip in tips:
        ws_instructions.append([tip])
    
    # Save template
    output_path = "templates/radar_template.xlsx"
    wb.save(output_path)
    print(f"✅ Template created: {output_path}")

if __name__ == "__main__":
    create_template()

