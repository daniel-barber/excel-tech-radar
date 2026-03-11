#!/usr/bin/env python3
"""Add isStrategic column to Excel template."""

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

# Paths
template_path = Path("templates/radar_template.xlsx")
data_template_path = Path("data/radar_template.xlsx")

def update_template(file_path):
    """Add isStrategic column with data validation."""
    print(f"Updating {file_path}...")
    
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Find the column headers row (usually row 1)
    headers = [cell.value for cell in ws[1]]
    print(f"Current headers: {headers}")
    
    # Check if isStrategic column already exists
    if 'isStrategic' in headers:
        print("isStrategic column already exists")
        wb.close()
        return
    
    # Find the position to insert (after propensityToWin if it exists, otherwise at the end)
    if 'propensityToWin' in headers:
        insert_col = headers.index('propensityToWin') + 2  # +1 for index, +1 for after
    else:
        insert_col = len(headers) + 1
    
    # Insert column header
    ws.cell(row=1, column=insert_col, value='isStrategic')
    print(f"Added isStrategic column at position {insert_col}")
    
    # Add data validation for isStrategic (TRUE/FALSE)
    dv = DataValidation(
        type="list",
        formula1='"TRUE,FALSE"',
        allow_blank=True
    )
    dv.error = 'Please select TRUE or FALSE'
    dv.errorTitle = 'Invalid Value'
    
    # Apply validation to column (rows 2-1000)
    col_letter = ws.cell(row=1, column=insert_col).column_letter
    dv.add(f'{col_letter}2:{col_letter}1000')
    ws.add_data_validation(dv)
    print(f"Added data validation to column {col_letter}")
    
    # Save
    wb.save(file_path)
    wb.close()
    print(f"✓ Updated {file_path}")

# Update both templates
if template_path.exists():
    update_template(template_path)
else:
    print(f"Template not found: {template_path}")

if data_template_path.exists():
    update_template(data_template_path)
else:
    print(f"Template not found: {data_template_path}")

print("\nDone!")

# Made with Bob
