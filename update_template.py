#!/usr/bin/env python3
"""Script to add dealSize column to the Excel template."""

from pathlib import Path
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

def update_template():
    """Add dealSize column with data validation to the template."""
    template_path = Path("templates/radar_template.xlsx")
    
    if not template_path.exists():
        print(f"Error: Template not found at {template_path}")
        return
    
    # Load workbook
    wb = openpyxl.load_workbook(template_path)
    
    # Update Sheet1 (data sheet)
    if "Sheet1" in wb.sheetnames:
        ws = wb["Sheet1"]
        
        # Find the last column (should be linkName in column H)
        # Add dealSize in column I
        ws["I1"] = "dealSize"
        
        # Add data validation for dealSize column
        dv = DataValidation(
            type="list",
            formula1='"< $100k,$100k - $500k,> $500k"',
            allow_blank=True
        )
        dv.error = "Please select a valid deal size"
        dv.errorTitle = "Invalid Deal Size"
        dv.prompt = "Select deal size category"
        dv.promptTitle = "Deal Size"
        
        # Apply validation to column I (rows 2-1000)
        ws.add_data_validation(dv)
        dv.add("I2:I1000")
        
        print("✓ Added dealSize column to Sheet1")
    
    # Update Instructions sheet
    if "Instructions" in wb.sheetnames:
        ws_inst = wb["Instructions"]
        
        # Find the row with column descriptions (should be around row 8-15)
        # We'll add the dealSize row after the linkName row
        
        # Look for the linkName row
        for row_idx in range(1, 30):
            cell_value = ws_inst.cell(row=row_idx, column=1).value
            if cell_value and "linkName" in str(cell_value):
                # Insert new row after linkName
                insert_row = row_idx + 1
                ws_inst.insert_rows(insert_row)
                
                # Add dealSize documentation
                ws_inst.cell(row=insert_row, column=1, value="dealSize")
                ws_inst.cell(row=insert_row, column=2, value="No")
                ws_inst.cell(row=insert_row, column=3, value="Deal size category: < $100k, $100k - $500k, > $500k")
                ws_inst.cell(row=insert_row, column=4, value="$100k - $500k")
                
                print("✓ Updated Instructions sheet")
                break
        
        # Update the ring values in instructions to match new config
        for row_idx in range(1, 50):
            cell_value = ws_inst.cell(row=row_idx, column=3).value
            if cell_value and "Ready, Less Than 1 Year" in str(cell_value):
                ws_inst.cell(row=row_idx, column=3, value="Time horizon: Q1, Q2, Q3, Q4, Beyond This Year")
                print("✓ Updated ring values in Instructions")
                break
        
        for row_idx in range(1, 50):
            cell_value = ws_inst.cell(row=row_idx, column=4).value
            if cell_value and "Ready" in str(cell_value):
                ws_inst.cell(row=row_idx, column=4, value="Q1")
                print("✓ Updated ring example in Instructions")
                break
        
        for row_idx in range(1, 50):
            cell_value = ws_inst.cell(row=row_idx, column=3).value
            if cell_value and "Valid rings:" in str(cell_value):
                ws_inst.cell(row=row_idx, column=3, value="Ring values must match: Q1, Q2, Q3, Q4, or Beyond This Year")
                print("✓ Updated ring validation note in Instructions")
                break
    
    # Save the updated template
    wb.save(template_path)
    print(f"\n✓ Template updated successfully: {template_path}")
    print("\nChanges made:")
    print("  - Added 'dealSize' column (column I) to Sheet1")
    print("  - Added data validation dropdown with three options:")
    print("    • < $100k")
    print("    • $100k - $500k")
    print("    • > $500k")
    print("  - Updated Instructions sheet with dealSize documentation")
    print("  - Updated ring values to Q1, Q2, Q3, Q4, Beyond This Year")

if __name__ == "__main__":
    update_template()

# Made with Bob
