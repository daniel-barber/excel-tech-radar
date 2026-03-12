"""
Flask API server for Excel Tech Radar management.
Provides REST endpoints for managing radar projects and entries.
"""
import json
import re
import shutil
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, Any
from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
import pandas as pd
from openpyxl import Workbook, load_workbook

from .loader import auto_discover_config, load_excel, validate_entries
from .builder import build_radar_json
from .config import Config
from .logging_config import setup_logging, get_logger
from .backup_manager import BackupManager
from .scheduler import create_default_tasks, TaskScheduler


class RadarAPI:
    """Flask API for radar management."""
    
    def __init__(self, config: Optional[Config] = None):
        """
        Initialize the Radar API.
        
        Args:
            config: Configuration object. If None, loads from environment.
        """
        if config is None:
            from .config import load_config
            config = load_config()
        
        self.config = config
        self.app = Flask(__name__)
        
        # Set up logging
        self.logger = setup_logging(
            level=config.log_level,
            log_file=config.log_file if config.log_file else None,
            log_format=config.log_format,
        )
        self.logger.info("Initializing Radar API")
        
        # Configure CORS
        cors_origins = config.allowed_origins
        if cors_origins == '*':
            CORS(self.app)
            self.logger.warning("CORS enabled for all origins - not recommended for production")
        else:
            CORS(self.app, origins=cors_origins.split(','))
            self.logger.info(f"CORS enabled for origins: {cors_origins}")
        
        # Set up directories
        self.data_dir = config.data_dir
        self.dist_dir = config.dist_dir
        self.max_backups = config.max_backups
        self.data_dir.mkdir(exist_ok=True)
        self.dist_dir.mkdir(exist_ok=True)
        self.logger.info(f"Data directory: {self.data_dir.absolute()}")
        self.logger.info(f"Dist directory: {self.dist_dir.absolute()}")
        
        # Initialize backup manager
        self.backup_manager = BackupManager(
            data_dir=self.data_dir,
            max_backups=self.max_backups,
            retention_days=config.retention_days
        )
        self.logger.info("Backup manager initialized")
        
        # Initialize task scheduler
        self.scheduler = create_default_tasks(self.backup_manager, config)
        if config.enable_scheduler if hasattr(config, 'enable_scheduler') else True:
            self.scheduler.start()
            self.logger.info("Task scheduler started")
        else:
            self.logger.info("Task scheduler disabled")
        
        # Register error handlers
        self._register_error_handlers()
        
        self._register_routes()
        self.logger.info("Radar API initialized successfully")
    
    def _cleanup_old_backups(self, project_id: str):
        """
        Keep only the N most recent backup files for a project.
        Removes .bak files older than the max_backups limit.
        """
        try:
            # Find all backup files for this project
            backup_pattern = f"{project_id}.xlsx.bak*"
            backup_files = list(self.data_dir.glob(backup_pattern))
            
            if len(backup_files) <= self.max_backups:
                return  # Nothing to clean up
            
            # Sort by modification time (newest first)
            backup_files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
            
            # Remove old backups beyond the limit
            for old_backup in backup_files[self.max_backups:]:
                old_backup.unlink()
                self.logger.info(f"Cleaned up old backup: {old_backup.name}", extra={'project_id': project_id})
        except Exception as e:
            self.logger.error(f"Error cleaning up backups for {project_id}: {e}", exc_info=True, extra={'project_id': project_id})
    
    def _cleanup_deleted_files(self, retention_days: int = 30):
        """
        Remove .deleted files older than retention_days.
        """
        try:
            from datetime import datetime, timedelta
            cutoff_time = datetime.now().timestamp() - (retention_days * 24 * 60 * 60)
            
            deleted_files = list(self.data_dir.glob("*.deleted"))
            cleaned_count = 0
            for deleted_file in deleted_files:
                if deleted_file.stat().st_mtime < cutoff_time:
                    deleted_file.unlink()
                    cleaned_count += 1
                    self.logger.debug(f"Cleaned up old deleted file: {deleted_file.name}")
            
            if cleaned_count > 0:
                self.logger.info(f"Cleaned up {cleaned_count} old deleted files")
        except Exception as e:
            self.logger.error(f"Error cleaning up deleted files: {e}", exc_info=True)
    
    def _build_radar_for_project(self, project_id: str) -> Dict[str, Any]:
        """Helper to build radar JSON for a project."""
        excel_file = self.data_dir / f"{project_id}.xlsx"
        
        # Determine sheet name
        try:
            xls = pd.ExcelFile(excel_file)
            sheet_name = 'Radar' if 'Radar' in xls.sheet_names else xls.sheet_names[0]
        except:
            sheet_name = 'Sheet1'
        
        # Auto-discover config
        config = auto_discover_config(
            excel_path=excel_file,
            sheet_name=sheet_name,
            title=project_id.replace('_', ' ')
        )
        
        # Load and validate entries
        raw_entries = load_excel(excel_file, sheet_name, config, allow_duplicates=False)
        entries = validate_entries(raw_entries)
        
        # Build radar JSON (to temp path, we'll return the dict)
        temp_json = self.dist_dir / 'temp_radar.json'
        radar_data = build_radar_json(config, entries, temp_json)
        
        return radar_data
    def _register_error_handlers(self):
        """Register error handlers for common HTTP errors."""
        
        @self.app.errorhandler(404)
        def not_found(error):
            """Handle 404 errors."""
            self.logger.warning(f"404 Not Found: {request.url}")
            return jsonify({
                'error': 'Resource not found',
                'status': 404,
                'path': request.path
            }), 404
        
        @self.app.errorhandler(500)
        def internal_error(error):
            """Handle 500 errors."""
            self.logger.error(f"500 Internal Server Error: {str(error)}", exc_info=True)
            return jsonify({
                'error': 'Internal server error',
                'status': 500,
                'message': str(error) if self.config.debug else 'An unexpected error occurred'
            }), 500
        
        @self.app.errorhandler(400)
        def bad_request(error):
            """Handle 400 errors."""
            self.logger.warning(f"400 Bad Request: {str(error)}")
            return jsonify({
                'error': 'Bad request',
                'status': 400,
                'message': str(error)
            }), 400
        
        @self.app.errorhandler(Exception)
        def handle_exception(error):
            """Handle uncaught exceptions."""
            self.logger.error(f"Uncaught exception: {str(error)}", exc_info=True)
            
            # Return JSON instead of HTML for HTTP errors
            if hasattr(error, 'code'):
                return jsonify({
                    'error': error.name if hasattr(error, 'name') else 'Error',
                    'status': error.code,
                    'message': str(error)
                }), error.code
            
            # Return 500 for other exceptions
            return jsonify({
                'error': 'Internal server error',
                'status': 500,
                'message': str(error) if self.config.debug else 'An unexpected error occurred'
            }), 500
    
    
    def _register_routes(self):
        """Register all API routes."""
        @self.app.route('/api/health', methods=['GET'])
        def health_check():
            """Health check endpoint for monitoring."""
            try:
                import psutil
                import time
                
                # Get system info
                disk_usage = psutil.disk_usage(str(self.data_dir))
                memory = psutil.virtual_memory()
                
                # Count projects
                project_count = len(list(self.data_dir.glob("*.xlsx")))
                
                health_data = {
                    'status': 'healthy',
                    'timestamp': datetime.utcnow().isoformat() + 'Z',
                    'version': '1.0.0',
                    'uptime_seconds': int(time.time() - self.app.config.get('START_TIME', time.time())),
                    'projects_count': project_count,
                    'disk': {
                        'total_gb': round(disk_usage.total / (1024**3), 2),
                        'used_gb': round(disk_usage.used / (1024**3), 2),
                        'free_gb': round(disk_usage.free / (1024**3), 2),
                        'percent_used': disk_usage.percent,
                    },
                    'memory': {
                        'total_gb': round(memory.total / (1024**3), 2),
                        'available_gb': round(memory.available / (1024**3), 2),
                        'percent_used': memory.percent,
                    },
                    'config': {
                        'data_dir': str(self.data_dir),
                        'max_backups': self.max_backups,
                        'debug': self.config.debug,
                    }
                }
                
                self.logger.debug("Health check requested")
                return jsonify(health_data)
            except ImportError:
                # psutil not installed, return basic health
                project_count = len(list(self.data_dir.glob("*.xlsx")))
                return jsonify({
                    'status': 'healthy',
                    'timestamp': datetime.utcnow().isoformat() + 'Z',
                    'projects_count': project_count,
                    'message': 'Install psutil for detailed system metrics'
                })
            except Exception as e:
                self.logger.error(f"Health check failed: {e}", exc_info=True)
                return jsonify({
                    'status': 'unhealthy',
                    'error': str(e)
                }), 500
        
        
        @self.app.route('/api/projects', methods=['GET'])
        def list_projects():
            """List all radar projects (Excel files in data/)."""
            try:
                projects = []
                for excel_file in self.data_dir.glob('*.xlsx'):
                    if excel_file.name.startswith('~$'):  # Skip temp files
                        continue
                    
                    stat = excel_file.stat()
                    
                    # Try to read display name from Excel properties
                    display_name = None
                    try:
                        wb = load_workbook(excel_file, read_only=True)
                        if wb.properties and wb.properties.title:
                            display_name = wb.properties.title
                        wb.close()
                    except Exception as e:
                        pass
                    
                    # Fallback to filename-based name if no stored display name
                    if not display_name:
                        # Use filename directly, just replace underscores with spaces
                        # Don't apply .title() as it would change "Stadt-ZH" to "Stadt-Zh"
                        display_name = excel_file.stem.replace('_', ' ')
                    
                    projects.append({
                        'id': excel_file.stem,
                        'name': display_name,
                        'filename': excel_file.name,
                        'size': stat.st_size,
                        'modified': datetime.fromtimestamp(stat.st_mtime).isoformat(),
                    })
                
                # Sort by modified date (newest first)
                projects.sort(key=lambda x: x['modified'], reverse=True)
                
                return jsonify({'projects': projects})
            except Exception as e:
                return jsonify({'error': str(e)}), 500
        
        @self.app.route('/api/projects/<project_id>', methods=['GET'])
        def get_project(project_id: str):
            """Get project details and radar data."""
            try:
                excel_file = self.data_dir / f"{project_id}.xlsx"
                if not excel_file.exists():
                    return jsonify({'error': 'Project not found'}), 404
                
                radar_data = self._build_radar_for_project(project_id)
                return jsonify(radar_data)
            except Exception as e:
                return jsonify({'error': str(e)}), 500
        
        @self.app.route('/api/projects/<project_id>/rename', methods=['POST'])
        def rename_project(project_id: str):
            """Rename a project (rename the Excel file)."""
            try:
                old_file = self.data_dir / f"{project_id}.xlsx"
                if not old_file.exists():
                    return jsonify({'error': 'Project not found'}), 404
                
                data = request.json
                new_name = data.get('new_name')
                
                if not new_name:
                    return jsonify({'error': 'new_name is required'}), 400
                
                # Validate filename (no special characters except underscore, hyphen, and space)
                # Remove .xlsx extension for validation if present
                name_to_validate = new_name.replace('.xlsx', '')
                if not re.match(r'^[a-zA-Z0-9_\- ]+$', name_to_validate):
                    return jsonify({'error': 'Project name can only contain letters, numbers, spaces, underscores, and hyphens'}), 400
                
                # Ensure .xlsx extension
                if not new_name.endswith('.xlsx'):
                    new_name += '.xlsx'
                
                new_file = self.data_dir / new_name
                
                # Check if target already exists
                if new_file.exists():
                    return jsonify({'error': 'A project with that name already exists'}), 400
                
                # Rename the file
                old_file.rename(new_file)
                
                return jsonify({
                    'success': True,
                    'old_name': f"{project_id}.xlsx",
                    'new_name': new_name
                })
            except Exception as e:
                return jsonify({'error': str(e)}), 500
        
        @self.app.route('/api/projects/<project_id>/excel', methods=['GET'])
        def get_excel_data(project_id: str):
            """Get raw Excel data as JSON for editing."""
            try:
                excel_file = self.data_dir / f"{project_id}.xlsx"
                if not excel_file.exists():
                    return jsonify({'error': 'Project not found'}), 404
                
                # Read Excel - try to get first sheet or 'Radar' sheet
                try:
                    # Try 'Radar' sheet first (our standard)
                    df = pd.read_excel(excel_file, sheet_name='Radar')
                except:
                    try:
                        # Try first sheet
                        df = pd.read_excel(excel_file, sheet_name=0)
                    except:
                        # Try 'Sheet1' as fallback
                        df = pd.read_excel(excel_file, sheet_name='Sheet1')
                
                # Replace NaN/inf with None for valid JSON
                df = df.replace([float('nan'), float('inf'), float('-inf')], None)
                
                # Convert to records (list of dicts)
                records = df.to_dict('records')
                
                # Get column names
                columns = df.columns.tolist()
                
                # Use Flask's jsonify which handles None properly
                return jsonify({
                    'columns': columns,
                    'rows': records,
                    'rowCount': len(records)
                })
            except Exception as e:
                import traceback
                traceback.print_exc()
                return jsonify({'error': str(e)}), 500
        
        @self.app.route('/api/projects/<project_id>/excel', methods=['PUT'])
        def update_excel_data(project_id: str):
            """Update Excel data from JSON."""
            try:
                excel_file = self.data_dir / f"{project_id}.xlsx"
                if not excel_file.exists():
                    return jsonify({'error': 'Project not found'}), 404
                
                data = request.json
                rows = data.get('rows', [])
                
                if not rows:
                    return jsonify({'error': 'No rows provided'}), 400
                
                # Determine sheet name from original file
                try:
                    xls = pd.ExcelFile(excel_file)
                    sheet_name = 'Radar' if 'Radar' in xls.sheet_names else xls.sheet_names[0]
                except:
                    sheet_name = 'Sheet1'
                
                # Create DataFrame from rows
                df = pd.DataFrame(rows)
                
                # Backup original file with timestamp
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                backup_file = excel_file.with_suffix(f'.xlsx.bak.{timestamp}')
                shutil.copy2(excel_file, backup_file)
                
                # Cleanup old backups
                self._cleanup_old_backups(project_id)
                
                # Write to Excel
                df.to_excel(excel_file, sheet_name=sheet_name, index=False)
                
                # Rebuild radar
                radar_data = self._build_radar_for_project(project_id)
                
                # Write radar.json to dist
                radar_json_file = self.dist_dir / 'radar.json'
                with open(radar_json_file, 'w') as f:
                    json.dump(radar_data, f, indent=2)
                
                return jsonify({
                    'success': True,
                    'message': f'Updated {len(rows)} rows',
                    'radar': radar_data
                })
            except Exception as e:
                return jsonify({'error': str(e)}), 500
        
        @self.app.route('/api/projects/<project_id>/rows', methods=['POST'])
        def add_row(project_id: str):
            """Add a new row to Excel."""
            try:
                excel_file = self.data_dir / f"{project_id}.xlsx"
                if not excel_file.exists():
                    return jsonify({'error': 'Project not found'}), 404
                
                row_data = request.json
                
                # Determine sheet name
                try:
                    xls = pd.ExcelFile(excel_file)
                    sheet_name = 'Radar' if 'Radar' in xls.sheet_names else xls.sheet_names[0]
                except:
                    sheet_name = 'Sheet1'
                
                # Read existing data
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                
                # Ensure all required columns exist
                required_cols = ['name', 'ring', 'quadrant', 'dealSize', 'propensityToWin', 'isStrategic', 'description', 'tags', 'link', 'linkName']
                for col in required_cols:
                    if col not in df.columns:
                        df[col] = pd.Series(dtype='object')
                
                # Append new row
                new_row = pd.DataFrame([row_data])
                df = pd.concat([df, new_row], ignore_index=True)
                
                # Write back
                df.to_excel(excel_file, sheet_name=sheet_name, index=False)
                
                # Rebuild radar
                radar_data = self._build_radar_for_project(project_id)
                
                # Write radar.json to dist
                radar_json_file = self.dist_dir / 'radar.json'
                with open(radar_json_file, 'w') as f:
                    json.dump(radar_data, f, indent=2)
                
                return jsonify({
                    'success': True,
                    'message': 'Row added',
                    'radar': radar_data
                })
            except Exception as e:
                return jsonify({'error': str(e)}), 500
        
        @self.app.route('/api/projects/<project_id>/rows/<int:row_index>', methods=['DELETE'])
        def delete_row(project_id: str, row_index: int):
            """Delete a row from Excel."""
            try:
                excel_file = self.data_dir / f"{project_id}.xlsx"
                if not excel_file.exists():
                    return jsonify({'error': 'Project not found'}), 404
                
                # Determine sheet name
                try:
                    xls = pd.ExcelFile(excel_file)
                    sheet_name = 'Radar' if 'Radar' in xls.sheet_names else xls.sheet_names[0]
                except:
                    sheet_name = 'Sheet1'
                
                # Read existing data
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                
                if row_index < 0 or row_index >= len(df):
                    return jsonify({'error': 'Invalid row index'}), 400
                
                # Drop row
                df = df.drop(row_index).reset_index(drop=True)
                
                # Write back
                df.to_excel(excel_file, sheet_name=sheet_name, index=False)
                
                # Rebuild radar and return updated data
                radar_data = self._build_radar_for_project(project_id)
                
                # Write radar.json to dist
                radar_json_file = self.dist_dir / 'radar.json'
                with open(radar_json_file, 'w') as f:
                    json.dump(radar_data, f, indent=2)
                
                return jsonify({
                    'success': True,
                    'message': 'Row deleted',
                    'radar': radar_data
                })
            except Exception as e:
                return jsonify({'error': str(e)}), 500
        
        @self.app.route('/api/projects/<project_id>/rows/<int:row_index>', methods=['PUT'])
        def update_row(project_id: str, row_index: int):
            """Update a single row in Excel."""
            try:
                excel_file = self.data_dir / f"{project_id}.xlsx"
                if not excel_file.exists():
                    return jsonify({'error': 'File not found'}), 404
                
                row_data = request.json
                
                # Determine sheet name
                try:
                    xls = pd.ExcelFile(excel_file)
                    sheet_name = 'Radar' if 'Radar' in xls.sheet_names else xls.sheet_names[0]
                except:
                    sheet_name = 'Sheet1'
                
                # Read existing data
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                
                if row_index < 0 or row_index >= len(df):
                    return jsonify({'error': 'Invalid row index'}), 400
                
                # Update the row - handle data type conversions
                # Add missing columns if they don't exist
                for key in row_data.keys():
                    if key not in df.columns:
                        # Add column with empty string default and explicit string dtype
                        df[key] = pd.Series([''] * len(df), dtype='object')
                
                # First pass: convert column types if needed
                # Convert any numeric columns that should be strings
                string_fields = ['name', 'ring', 'quadrant', 'status', 'dealSize', 'propensityToWin', 'description', 'tags', 'link', 'linkName', 'customer', 'owner']
                for key in string_fields:
                    if key in df.columns and pd.api.types.is_numeric_dtype(df[key]):
                        df[key] = df[key].astype('object')
                
                # Also check incoming data for string values or booleans going to numeric columns
                for key, value in row_data.items():
                    if key in df.columns:
                        # If trying to set a string value (including empty string) to a numeric column, convert column to object first
                        if isinstance(value, str) and pd.api.types.is_numeric_dtype(df[key]):
                            df[key] = df[key].astype('object')
                        # If trying to set a boolean to a numeric column, convert column to object first
                        elif isinstance(value, bool) and pd.api.types.is_numeric_dtype(df[key]):
                            df[key] = df[key].astype('object')
                
                # Second pass: set values
                for key, value in row_data.items():
                    if key in df.columns:
                        # Convert tags array to comma-separated string
                        if key == 'tags' and isinstance(value, list):
                            value = ', '.join(value) if value else ''
                        # Convert boolean to proper format
                        elif key in ['isNew', 'isStrategic'] and isinstance(value, bool):
                            value = value
                        # Handle None/null/empty values based on column dtype
                        elif value is None or value == '':
                            # For numeric columns, use None (becomes NaN in pandas)
                            if pd.api.types.is_numeric_dtype(df[key]):
                                value = None
                            # For string columns, use empty string
                            else:
                                value = ''
                        
                        try:
                            df.at[row_index, key] = value
                        except Exception as e:
                            # Log the error with details
                            print(f"Error setting {key}={value} (type: {type(value)}, dtype: {df[key].dtype}): {e}")
                            raise ValueError(f"Invalid value '{value}' for dtype '{df[key].dtype}'")
                
                # Backup original file with timestamp
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                backup_file = excel_file.with_suffix(f'.xlsx.bak.{timestamp}')
                shutil.copy2(excel_file, backup_file)
                
                # Cleanup old backups
                self._cleanup_old_backups(project_id)
                
                # Write back
                df.to_excel(excel_file, sheet_name=sheet_name, index=False)
                
                # Rebuild radar
                radar_data = self._build_radar_for_project(project_id)
                
                return jsonify({
                    'success': True,
                    'message': 'Row updated',
                    'radar': radar_data
                })
            except Exception as e:
                import traceback
                traceback.print_exc()
                return jsonify({'error': str(e)}), 500
        
        @self.app.route('/api/projects', methods=['POST'])
        def create_project():
            """Create a new radar project."""
            try:
                data = request.json
                project_name = data.get('name', 'New Radar')
                template = data.get('template', 'default')
                
                # Sanitize project name for filename (allow alphanumeric, hyphens, and underscores)
                project_id = project_name.lower().replace(' ', '_')
                project_id = ''.join(c for c in project_id if c.isalnum() or c in ('_', '-'))
                
                excel_file = self.data_dir / f"{project_id}.xlsx"
                
                # Check if exists
                if excel_file.exists():
                    return jsonify({'error': 'Project already exists'}), 400
                
                # Create new Excel with template
                wb = Workbook()
                ws = wb.active
                ws.title = 'Sheet1'
                
                # Store the original display name in document properties
                wb.properties.title = project_name
                
                # Add headers based on template
                if template == 'default':
                    headers = ['name', 'ring', 'quadrant', 'status', 'description', 'tags', 'link', 'linkName']
                    ws.append(headers)
                    # No sample rows - start with empty project
                
                wb.save(excel_file)
                
                return jsonify({
                    'success': True,
                    'project': {
                        'id': project_id,
                        'name': project_name,
                        'filename': excel_file.name,
                    }
                })
            except Exception as e:
                return jsonify({'error': str(e)}), 500
        
        @self.app.route('/api/projects/<project_id>', methods=['DELETE'])
        def delete_project(project_id: str):
            """Delete a radar project."""
            try:
                excel_file = self.data_dir / f"{project_id}.xlsx"
                if not excel_file.exists():
                    return jsonify({'error': 'Project not found'}), 404
                
                # Move to trash (rename with .deleted suffix)
                deleted_file = excel_file.with_suffix('.xlsx.deleted')
                excel_file.rename(deleted_file)
                
                return jsonify({'success': True, 'message': 'Project deleted'})
            except Exception as e:
                return jsonify({'error': str(e)}), 500
        @self.app.route('/api/maintenance/cleanup', methods=['POST'])
        def cleanup_maintenance():
            """Manual cleanup of old backups and deleted files."""
            try:
                data = request.get_json() or {}
                retention_days = data.get('retention_days', 30)
                
                # Use backup manager for cleanup
                stats = self.backup_manager.cleanup_all_backups()
                
                return jsonify({
                    'success': True,
                    'message': 'Cleanup completed',
                    'stats': stats
                })
            except Exception as e:
                return jsonify({'error': str(e)}), 500
        
        @self.app.route('/api/backups', methods=['GET'])
        def list_all_backups():
            """List all backups across all projects."""
            try:
                project_id = request.args.get('project_id')
                backups = self.backup_manager.list_backups(project_id)
                
                return jsonify({
                    'backups': [b.to_dict() for b in backups],
                    'count': len(backups)
                })
            except Exception as e:
                self.logger.error(f"Failed to list backups: {e}", exc_info=True)
                return jsonify({'error': str(e)}), 500
        
        @self.app.route('/api/projects/<project_id>/backups', methods=['GET'])
        def list_project_backups(project_id: str):
            """List backups for a specific project."""
            try:
                backups = self.backup_manager.list_backups(project_id)
                
                return jsonify({
                    'project_id': project_id,
                    'backups': [b.to_dict() for b in backups],
                    'count': len(backups)
                })
            except Exception as e:
                self.logger.error(f"Failed to list backups for {project_id}: {e}", exc_info=True, extra={'project_id': project_id})
                return jsonify({'error': str(e)}), 500
        
        @self.app.route('/api/projects/<project_id>/backups', methods=['POST'])
        def create_backup(project_id: str):
            """Create a manual backup of a project."""
            try:
                excel_file = self.data_dir / f"{project_id}.xlsx"
                if not excel_file.exists():
                    return jsonify({'error': 'Project not found'}), 404
                
                backup_path = self.backup_manager.create_backup(project_id, backup_type='manual')
                
                if backup_path:
                    return jsonify({
                        'success': True,
                        'message': 'Backup created successfully',
                        'backup_file': backup_path.name,
                        'size_mb': round(backup_path.stat().st_size / (1024 * 1024), 2)
                    })
                else:
                    return jsonify({'error': 'Failed to create backup'}), 500
            except Exception as e:
                self.logger.error(f"Failed to create backup for {project_id}: {e}", exc_info=True, extra={'project_id': project_id})
                return jsonify({'error': str(e)}), 500
        
        @self.app.route('/api/projects/<project_id>/backups/<backup_timestamp>', methods=['DELETE'])
        def delete_backup(project_id: str, backup_timestamp: str):
            """Delete a specific backup."""
            try:
                success = self.backup_manager.delete_backup(project_id, backup_timestamp)
                
                if success:
                    return jsonify({
                        'success': True,
                        'message': 'Backup deleted successfully'
                    })
                else:
                    return jsonify({'error': 'Backup not found'}), 404
            except Exception as e:
                self.logger.error(f"Failed to delete backup: {e}", exc_info=True, extra={'project_id': project_id})
                return jsonify({'error': str(e)}), 500
        
        @self.app.route('/api/projects/<project_id>/restore', methods=['POST'])
        def restore_from_backup(project_id: str):
            """Restore a project from a backup."""
            try:
                data = request.get_json() or {}
                backup_timestamp = data.get('backup_timestamp')
                
                success = self.backup_manager.restore_backup(project_id, backup_timestamp)
                
                if success:
                    # Rebuild radar after restore
                    radar_data = self._build_radar_for_project(project_id)
                    
                    return jsonify({
                        'success': True,
                        'message': 'Project restored successfully',
                        'radar': radar_data
                    })
                else:
                    return jsonify({'error': 'Failed to restore backup'}), 500
            except Exception as e:
                self.logger.error(f"Failed to restore backup for {project_id}: {e}", exc_info=True, extra={'project_id': project_id})
                return jsonify({'error': str(e)}), 500
        
        @self.app.route('/api/projects/<project_id>/export', methods=['POST'])
        def export_project(project_id: str):
            """Export a project with all backups as a ZIP file."""
            try:
                excel_file = self.data_dir / f"{project_id}.xlsx"
                if not excel_file.exists():
                    return jsonify({'error': 'Project not found'}), 404
                
                export_path = self.backup_manager.export_project(project_id)
                
                if export_path:
                    return send_file(
                        export_path,
                        as_attachment=True,
                        download_name=export_path.name,
                        mimetype='application/zip'
                    )
                else:
                    return jsonify({'error': 'Failed to export project'}), 500
            except Exception as e:
                self.logger.error(f"Failed to export project {project_id}: {e}", exc_info=True, extra={'project_id': project_id})
                return jsonify({'error': str(e)}), 500
        
        @self.app.route('/api/projects/import', methods=['POST'])
        def import_project():
            """Import a project from a ZIP export file."""
            try:
                if 'file' not in request.files:
                    return jsonify({'error': 'No file provided'}), 400
                
                uploaded_file = request.files['file']
                
                if uploaded_file.filename == '':
                    return jsonify({'error': 'No file selected'}), 400
                
                if not uploaded_file.filename.endswith('.zip'):
                    return jsonify({'error': 'Invalid file type. Please upload a ZIP file'}), 400
                
                # Save uploaded file temporarily
                temp_path = self.data_dir / f"temp_import_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
                uploaded_file.save(temp_path)
                
                try:
                    # Check if overwrite is allowed
                    overwrite = request.form.get('overwrite', 'false').lower() == 'true'
                    
                    # Import project
                    project_id = self.backup_manager.import_project(temp_path, overwrite=overwrite)
                    
                    if project_id:
                        return jsonify({
                            'success': True,
                            'message': 'Project imported successfully',
                            'project_id': project_id
                        })
                    else:
                        return jsonify({'error': 'Failed to import project'}), 500
                finally:
                    # Clean up temp file
                    if temp_path.exists():
                        temp_path.unlink()
            except Exception as e:
                self.logger.error(f"Failed to import project: {e}", exc_info=True)
                return jsonify({'error': str(e)}), 500
        
        @self.app.route('/api/storage/stats', methods=['GET'])
        def get_storage_stats():
            """Get storage statistics for the data directory."""
            try:
                stats = self.backup_manager.get_storage_stats()
                return jsonify(stats)
            except Exception as e:
                self.logger.error(f"Failed to get storage stats: {e}", exc_info=True)
                return jsonify({'error': str(e)}), 500
        
        @self.app.route('/api/scheduler/status', methods=['GET'])
        def get_scheduler_status():
            """Get scheduler status and task information."""
            try:
                return jsonify({
                    'running': self.scheduler.is_running(),
                    'tasks': self.scheduler.get_task_status()
                })
            except Exception as e:
                self.logger.error(f"Failed to get scheduler status: {e}", exc_info=True)
                return jsonify({'error': str(e)}), 500
        
        @self.app.route('/api/scheduler/start', methods=['POST'])
        def start_scheduler():
            """Start the task scheduler."""
            try:
                if self.scheduler.is_running():
                    return jsonify({'message': 'Scheduler already running'}), 200
                
                self.scheduler.start()
                return jsonify({
                    'success': True,
                    'message': 'Scheduler started'
                })
            except Exception as e:
                self.logger.error(f"Failed to start scheduler: {e}", exc_info=True)
                return jsonify({'error': str(e)}), 500
        
        @self.app.route('/api/scheduler/stop', methods=['POST'])
        def stop_scheduler():
            """Stop the task scheduler."""
            try:
                if not self.scheduler.is_running():
                    return jsonify({'message': 'Scheduler not running'}), 200
                
                self.scheduler.stop()
                return jsonify({
                    'success': True,
                    'message': 'Scheduler stopped'
                })
            except Exception as e:
                self.logger.error(f"Failed to stop scheduler: {e}", exc_info=True)
                return jsonify({'error': str(e)}), 500
        
        @self.app.route('/api/scheduler/tasks/<task_name>/enable', methods=['POST'])
        def enable_task(task_name: str):
            """Enable a specific scheduled task."""
            try:
                success = self.scheduler.enable_task(task_name)
                if success:
                    return jsonify({
                        'success': True,
                        'message': f'Task "{task_name}" enabled'
                    })
                else:
                    return jsonify({'error': 'Task not found'}), 404
            except Exception as e:
                self.logger.error(f"Failed to enable task {task_name}: {e}", exc_info=True)
                return jsonify({'error': str(e)}), 500
        
        @self.app.route('/api/scheduler/tasks/<task_name>/disable', methods=['POST'])
        def disable_task(task_name: str):
            """Disable a specific scheduled task."""
            try:
                success = self.scheduler.disable_task(task_name)
                if success:
                    return jsonify({
                        'success': True,
                        'message': f'Task "{task_name}" disabled'
                    })
                else:
                    return jsonify({'error': 'Task not found'}), 404
            except Exception as e:
                self.logger.error(f"Failed to disable task {task_name}: {e}", exc_info=True)
                return jsonify({'error': str(e)}), 500
        
        
        @self.app.route('/api/projects/<project_id>/download', methods=['GET'])
        def download_project(project_id: str):
            """Download Excel file."""
            try:
                excel_file = self.data_dir / f"{project_id}.xlsx"
                if not excel_file.exists():
                    return jsonify({'error': 'Project not found'}), 404
                
                return send_file(
                    excel_file,
                    as_attachment=True,
                    download_name=f"{project_id}.xlsx"
                )
            except Exception as e:
                return jsonify({'error': str(e)}), 500
        
        @self.app.route('/api/projects/<project_id>/build', methods=['POST'])
        def build_project(project_id: str):
            """Build radar JSON for a project."""
            try:
                excel_file = self.data_dir / f"{project_id}.xlsx"
                if not excel_file.exists():
                    return jsonify({'error': 'Project not found'}), 404
                
                radar_data = self._build_radar_for_project(project_id)
                
                # Write to dist
                radar_json_file = self.dist_dir / 'radar.json'
                with open(radar_json_file, 'w') as f:
                    json.dump(radar_data, f, indent=2)
                
                return jsonify({
                    'success': True,
                    'message': 'Radar built successfully',
                    'radar': radar_data
                })
            except Exception as e:
                return jsonify({'error': str(e)}), 500
        
        @self.app.route('/api/template/download', methods=['GET'])
        def download_template():
            """Download Excel template for importing projects."""
            try:
                template_dir = Path(__file__).parent.parent.parent / 'templates'
                template_file = template_dir / 'radar_template.xlsx'
                
                if not template_file.exists():
                    return jsonify({'error': 'Template file not found'}), 404
                
                return send_file(
                    template_file,
                    as_attachment=True,
                    download_name='radar_template.xlsx',
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            except Exception as e:
                return jsonify({'error': str(e)}), 500
        
        @self.app.route('/api/projects/upload', methods=['POST'])
        def upload_project():
            """Upload an Excel file to create a new project."""
            try:
                # Check if file is present
                if 'file' not in request.files:
                    return jsonify({'error': 'No file provided'}), 400
                
                uploaded_file = request.files['file']
                
                if uploaded_file.filename == '':
                    return jsonify({'error': 'No file selected'}), 400
                
                # Validate file extension
                if not (uploaded_file.filename.endswith('.xlsx') or uploaded_file.filename.endswith('.xls')):
                    return jsonify({'error': 'Invalid file type. Please upload an Excel file (.xlsx or .xls)'}), 400
                
                # Generate project ID from filename
                filename = uploaded_file.filename.rsplit('.', 1)[0]  # Remove extension
                project_id = re.sub(r'[^a-z0-9_-]+', '_', filename.lower())
                
                # Check if project already exists
                excel_path = self.data_dir / f"{project_id}.xlsx"
                if excel_path.exists():
                    return jsonify({'error': f'Project "{filename}" already exists'}), 409
                
                # Save the file
                uploaded_file.save(str(excel_path))
                
                # Validate the file can be loaded and set document properties in one operation
                try:
                    # First validate by building radar data (this loads the file)
                    radar_data = self._build_radar_for_project(project_id)
                    
                    # Then set document properties with original filename
                    wb = load_workbook(excel_path)
                    wb.properties.title = filename
                    wb.save(excel_path)
                except Exception as e:
                    # Delete the file if validation fails
                    excel_path.unlink()
                    return jsonify({'error': f'Invalid Excel file: {str(e)}'}), 400
                
                return jsonify({
                    'success': True,
                    'message': 'Project uploaded successfully',
                    'project_id': project_id,
                    'project_name': filename
                })
                
            except Exception as e:
                import traceback
                traceback.print_exc()
                return jsonify({'error': str(e)}), 500
        
        # Serve static files
        @self.app.route('/')
        def index():
            """Serve unified interface."""
            web_dir = Path(__file__).parent.parent.parent / 'web'
            return send_from_directory(str(web_dir), 'unified.html')
        
        @self.app.route('/<path:path>')
        def static_files(path):
            """Serve static files from web/ or dist/."""
            # Get absolute paths
            web_dir = Path(__file__).parent.parent.parent / 'web'
            dist_dir = Path(__file__).parent.parent.parent / 'dist'
            
            # Try web/ first, then dist/
            web_file = web_dir / path
            if web_file.exists():
                return send_from_directory(str(web_dir), path)
            
            dist_file = dist_dir / path
            if dist_file.exists():
                return send_from_directory(str(dist_dir), path)
            
            return jsonify({'error': 'File not found'}), 404
    
    def run(self, host='127.0.0.1', port=5173, debug=True):
        """Run the Flask server."""
        print(f"🚀 Starting Radar API server on http://{host}:{port}")
        print(f"📁 Data directory: {self.data_dir.absolute()}")
        print(f"📦 Dist directory: {self.dist_dir.absolute()}")
        self.app.run(host=host, port=port, debug=debug)


def create_api(data_dir: str = "data", dist_dir: str = "dist") -> Flask:
    """Create and return Flask app instance."""
    api = RadarAPI(data_dir=data_dir, dist_dir=dist_dir)
    return api.app

# Made with Bob
