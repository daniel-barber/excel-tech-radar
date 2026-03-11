#!/usr/bin/env python3
"""Add propensityToWin column to Excel template."""

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

# Paths
template_path = Path("templates/radar_template.xlsx")
data_template_path = Path("data/radar_template.xlsx")

def update_template(file_path):
    """Add propensityToWin column with data validation."""
    print(f"Updating {file_path}...")
    
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Find the column headers row (usually row 1)
    headers = [cell.value for cell in ws[1]]
    print(f"Current headers: {headers}")
    
    # Check if propensityToWin column already exists
    if 'propensityToWin' in headers:
        print("propensityToWin column already exists")
        wb.close()
        return
    
    # Find the position to insert (after dealSize if it exists, otherwise at the end)
    if 'dealSize' in headers:
        insert_col = headers.index('dealSize') + 2  # +1 for index, +1 for after
    else:
        insert_col = len(headers) + 1
    
    # Insert column header
    ws.cell(row=1, column=insert_col, value='propensityToWin')
    print(f"Added propensityToWin column at position {insert_col}")
    
    # Add data validation for propensityToWin
    dv = DataValidation(
        type="list",
        formula1='"Low,Medium,High"',
        allow_blank=True
    )
    dv.error = 'Please select from the list'
    dv.errorTitle = 'Invalid Propensity'
    
    # Apply validation to column (rows 2-1000)
    col_letter = ws.cell(row=1, column=insert_col).column_letter
    dv.add(f'{col_letter}2:{col_letter}1000')
    ws.add_data_validation(dv)
    print(f"Added data validation to column {col_letter}")
    
    # Save
    wb.save(file_path)
    wb.close()
    print(f"✓ Updated {file_path}")

# Update both templates
if template_path.exists():
    update_template(template_path)
else:
    print(f"Template not found: {template_path}")

if data_template_path.exists():
    update_template(data_template_path)
else:
    print(f"Template not found: {data_template_path}")

print("\nDone!")

