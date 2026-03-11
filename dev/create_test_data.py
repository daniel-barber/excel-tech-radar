#!/usr/bin/env python3
"""Create test Excel file with dealSize data."""

from pathlib import Path
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

def create_test_data():
    """Create test Excel file with sample data including dealSize."""
    
    # Create new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Radar"
    
    # Add headers
    headers = ["name", "ring", "quadrant", "status", "dealSize", "description", "tags", "link", "linkName"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Add sample data with different deal sizes
    sample_data = [
        ["Cloud Migration Project", "Q1", "Infrastructure", "On Track", "< $100k", 
         "<p>Small cloud migration initiative</p>", "cloud, migration", "", ""],
        ["Enterprise CRM System", "Q1", "Applications", "On Track", "$100k - $500k",
         "<p>Medium-sized CRM implementation</p>", "crm, sales", "", ""],
        ["Digital Transformation Program", "Q2", "Services", "At Risk", "> $500k",
         "<p>Large-scale digital transformation</p>", "transformation, strategy", "", ""],
        ["AI Analytics Platform", "Q2", "Emerging Tech", "New", "$100k - $500k",
         "<p>AI-powered analytics solution</p>", "ai, analytics", "", ""],
        ["Security Upgrade", "Q3", "Infrastructure", "On Track", "< $100k",
         "<p>Security infrastructure upgrade</p>", "security, infrastructure", "", ""],
        ["Data Warehouse Modernization", "Q3", "Applications", "Blocked", "> $500k",
         "<p>Large data warehouse project</p>", "data, warehouse", "", ""],
        ["IoT Platform", "Q4", "Emerging Tech", "New", "$100k - $500k",
         "<p>IoT platform development</p>", "iot, platform", "", ""],
        ["Legacy System Replacement", "Beyond This Year", "Applications", "On Track", "> $500k",
         "<p>Long-term legacy replacement</p>", "legacy, modernization", "", ""],
    ]
    
    # Add data rows
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Add data validation for dealSize column (column E)
    dv = DataValidation(
        type="list",
        formula1='"< $100k,$100k - $500k,> $500k"',
        allow_blank=True
    )
    dv.error = "Please select a valid deal size"
    dv.errorTitle = "Invalid Deal Size"
    ws.add_data_validation(dv)
    dv.add("E2:E1000")
    
    # Save file
    output_path = Path("data/test_dealsize.xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    
    print(f"✓ Created test file: {output_path}")
    print("\nSample data includes:")
    print("  - 8 entries across all rings (Q1, Q2, Q3, Q4, Beyond This Year)")
    print("  - Mix of deal sizes:")
    print("    • 2 small deals (< $100k)")
    print("    • 3 medium deals ($100k - $500k)")
    print("    • 3 large deals (> $500k)")
    print("  - Various statuses and quadrants")

if __name__ == "__main__":
    create_test_data()

